from flask import Flask, jsonify, render_template, request, redirect, url_for
import openpyxl
import smtplib
from datetime import datetime
from apscheduler.schedulers.background import BackgroundScheduler
import logging
import os

app = Flask(__name__)

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Email configuration
EMAIL_ADDRESS = 'anjalisingh88098@gmail.com'  # Your email address
EMAIL_PASSWORD = 'fzeuuncebuvybwgq'  # Your app-specific password

# Email sending function
def send_reminder_email(to_email, subject, body):
    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            message = f'Content-Type: text/html\nSubject: {subject}\n\n{body}'
            server.sendmail(EMAIL_ADDRESS, to_email, message)
        logging.info(f"Email sent to {to_email}")
    except Exception as e:
        logging.error(f"Failed to send email: {e}")

# Load Excel data and return it as a list of dictionaries
def load_excel_data():
    try:
        # Set the correct path for your Excel file
        workbook = openpyxl.load_workbook('D:/Lubrication Reminder System/lubrication_schedule.xlsx')
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]
        data = [dict(zip(headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True) if any(row)]
        
        logging.info(f"Loaded {len(data)} rows from Excel.")
        return data
    except Exception as e:
        logging.error(f"Error loading Excel data: {e}")
        return []

# Function to update the Excel sheet
def update_excel_row(serial_no, status, user_name=None, reason=None, suggested_time=None):
    try:
        workbook = openpyxl.load_workbook('D:/Lubrication Reminder System/lubrication_schedule.xlsx')
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2):
            if row[0].value == serial_no:  # Assuming SR.NO is in the first column
                if user_name:
                    row[6].value = user_name  # Assuming column 7 is for user name
                row[7].value = status  # Assuming column 8 is for status (Approved/Rejected)
                if reason:
                    row[8].value = reason  # Assuming column 9 is for rejection reason
                if suggested_time:
                    row[9].value = suggested_time  # Assuming column 10 is for suggested time
                break
        
        workbook.save('D:/Lubrication Reminder System/lubrication_schedule.xlsx')
        logging.info(f"Updated Excel row for serial number: {serial_no}")
    except Exception as e:
        logging.error(f"Error updating Excel row: {e}")

# Function to send a test email
def send_test_email():
    data = load_excel_data()
    if not data:
        logging.warning("No data found or error occurred in loading Excel file.")
        return

    row = data[0]
    serial_no = row.get('SR.NO', 'N/A')
    component = row.get('UNIT/ COMPONENTS', 'N/A')
    lubricant = row.get('LUBRICANTS', 'N/A')
    interval = row.get('INTERVAL', 'N/A')
    quantity = row.get('QTY', 'N/A')
    due_date = row.get('DUE DATE', 'N/A')

    # Check if the due date is valid and format it
    try:
        due_date_obj = datetime.strptime(due_date, '%d-%m-%Y')
        formatted_due_date = due_date_obj.strftime('%d-%m-%Y')
    except ValueError:
        formatted_due_date = 'Invalid Date'

    subject = f"Test Email: Task {component} Details"

    # HTML content for the test email
    body = f"""
    <html>
    <body>
        <h3>{component} Task Due Today.</h3>
        <p>This is a reminder for the pending task that needs your attention. Please review the task and approve it by clicking the button below.</p>

        <b>Task Details:</b>
        <ul>
            <li><strong>Serial Number:</strong> {serial_no}</li>
            <li><strong>Component:</strong> {component}</li>
            <li><strong>Lubricant:</strong> {lubricant}</li>
            <li><strong>Interval (Months):</strong> {interval}</li>
            <li><strong>Quantity:</strong> {quantity}</li>
            <li><strong>Due Date:</strong> {formatted_due_date}</li>
        </ul>

        <p>Please complete this task today to ensure the equipment continues running smoothly.</p>
        <p>
              <a href="{url_for('approve_task', serial_no=serial_no, _external=True)}" 
           style="background-color:green;color:white;padding:10px 15px;text-decoration:none;">Approve</a>
        <a href="{url_for('reject_task', serial_no=serial_no, _external=True)}" 
           style="background-color:red;color:white;padding:10px 15px;text-decoration:none;margin-left:10px;">Reject</a>
        </p>
        <p> Once the task is approved, the status will be updated in the system automatically.</p>
    </body>
    </html>
    """

    # Send the test email
    send_reminder_email(EMAIL_ADDRESS, subject, body)

# Route to send a test email
@app.route('/send_test_email', methods=['GET'])
def test_email_route():
    send_test_email()
    return "Test email sent!"

# Function to check due dates and send reminders
def check_due_dates():
    data = load_excel_data()
    if not data:
        logging.warning("No data found or error occurred in loading Excel file.")
        return

    today = datetime.now().date()

    for row in data:
        due_date = row.get('DUE DATE')
        if due_date:
            try:
                # Parse the due date from the format 'd-m-yyyy'
                due_date_obj = datetime.strptime(due_date, '%d-%m-%Y').date()
            except ValueError as e:
                logging.error(f"Error parsing date for task {row.get('UNIT/ COMPONENTS', 'Unknown')}: {due_date} - {e}")
                continue

            # Check if the due date matches today's date
            if due_date_obj == today:
                serial_no = row.get('SR.NO', 'N/A')
                component = row.get('UNIT/ COMPONENTS', 'N/A')
                lubricant = row.get('LUBRICANTS', 'N/A')
                interval = row.get('INTERVAL', 'N/A')
                quantity = row.get('QTY', 'N/A')

                subject = f"Reminder: Task {component} is due today!"

                # HTML content with Approve and Reject buttons
                body = f"""
                <html>
                <body>
                    <h3>Task {component} Due Today!</h3>
                    <p>Task Details:</p>
                    <ul>
                        <li><strong>Serial Number:</strong> {serial_no}</li>
                        <li><strong>Component:</strong> {component}</li>
                        <li><strong>Lubricant:</strong> {lubricant}</li>
                        <li><strong>Interval (Months):</strong> {interval}</li>
                        <li><strong>Quantity:</strong> {quantity}</li>
                        <li><strong>Due Date:</strong> {due_date_obj.strftime('%d-%m-%Y')}</li>
                    </ul>
                    <p>Please ensure this task is completed by today.</p>
                    <p>
                        <a href="{url_for('approve_task', serial_no=serial_no)}" style="background-color:green;color:white;padding:10px 15px;text-decoration:none;">Approve</a>
                        <a href="{url_for('reject_task', serial_no=serial_no)}" style="background-color:red;color:white;padding:10px 15px;text-decoration:none;margin-left:10px;">Reject</a>
                    </p>
                </body>
                </html>
                """
                # Send the email
                send_reminder_email(EMAIL_ADDRESS, subject, body)

# Scheduler for daily job
scheduler = BackgroundScheduler()
scheduler.add_job(func=check_due_dates, trigger="cron", hour=8, minute=0)
scheduler.start()

# Serve the frontend
@app.route('/')
def index():
    return render_template('index.html')

# Route to fetch data from Excel and display in JSON format
@app.route('/fetch_excel_data', methods=['GET'])
def fetch_excel_data():
    data = load_excel_data()
    if data:
        return jsonify(data)
    else:
        return jsonify({"message": "No data found or an error occurred."}), 404

@app.route('/approve/<serial_no>', methods=['GET'])
def approve_task(serial_no):
    user_name = EMAIL_ADDRESS.split('@')[0]  # Simulating user approval
    update_excel_row(serial_no, 'Approved', user_name)
    return redirect(url_for('index'))

@app.route('/reject/<serial_no>', methods=['GET'])
def reject_task(serial_no):
    user_name = EMAIL_ADDRESS.split('@')[0]  # Simulating user rejection
    return render_template('reject_task.html', serial_no=serial_no, user_name=user_name)

@app.route('/submit_rejection', methods=['POST'])
def submit_rejection():
    serial_no = request.form['serial_no']
    reason = request.form['reason']
    update_excel_row(serial_no, 'Rejected', reason=reason)
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
